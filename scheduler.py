
from collections import defaultdict
from copy import copy
from datetime import datetime
from itertools import permutations, product
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SLOT_MINUTES = 5
LUNCH_MINUTES = 30
LUNCH_WINDOW_START_MINUTE = 12 * 60
LUNCH_WINDOW_END_MINUTE = 16 * 60
OUTPUT_FILE = "DC_Schedule.xlsx"
FONT_NAME = "Montserrat"


ASSESSOR_COLORS = [
    "FFE070", "9DC3E6", "A9D18E", "C6A0DC",
    "F4B183", "B7DEE8", "D9EAD3", "F8CBAD",
]


def round_up_to_slots(minutes):
    if minutes <= 0:
        return 0
    return int(math.ceil(minutes / SLOT_MINUTES))


def slots_to_minutes(slots):
    return slots * SLOT_MINUTES


def rounded_duration(minutes):
    slots = round_up_to_slots(minutes)
    return slots, slots_to_minutes(slots)


def minutes_from_time(value):
    return value.hour * 60 + value.minute


def read_positive_int(prompt):
    while True:
        try:
            number = int(input(prompt).strip())
            if number > 0:
                return number
        except ValueError:
            pass
        print("Please enter a positive whole number.")


def read_minimum_int(prompt, minimum):
    while True:
        try:
            number = int(input(prompt).strip())
            if number >= minimum:
                return number
        except ValueError:
            pass
        print(f"Please enter a whole number of at least {minimum}.")


def read_non_negative_int(prompt):
    while True:
        try:
            number = int(input(prompt).strip())
            if number >= 0:
                return number
        except ValueError:
            pass
        print("Please enter zero or a positive whole number.")


def read_time(prompt):
    while True:
        value = input(prompt).strip()
        try:
            return datetime.strptime(value, "%H:%M")
        except ValueError:
            print("Please enter time in HH:MM format, for example 09:30.")


def read_end_time(start_time):
    start_minute = minutes_from_time(start_time)
    while True:
        end_time = read_time("DC end time (HH:MM): ")
        if minutes_from_time(end_time) > start_minute:
            return end_time
        print("End time must be later than the start time.")


def read_yes_no(prompt):
    while True:
        value = input(prompt).strip().lower()
        if value in ("y", "yes"):
            return True
        if value in ("n", "no"):
            return False
        print("Please enter y or n.")


def collect_optional_names(count, label, prefix):
    if not read_yes_no(f"Do you want to enter {label} names? (y/n): "):
        return [""] * count

    names = []
    for index in range(1, count + 1):
        names.append(input(f"Enter {label} name for {prefix}{index}: ").strip())
    return names


def lunch_slots():
    return round_up_to_slots(LUNCH_MINUTES)


def collect_inputs():
    print("Enter Development Center schedule details")
    candidates = read_positive_int("Number of candidates: ")
    assessors = read_minimum_int("Number of assessors: ", 1)
    start_time = read_time("Start time (HH:MM): ")
    end_time = read_end_time(start_time)
    context_minutes = read_non_negative_int("Context setting duration (minutes): ")
    tool_count = read_minimum_int("Number of tools: ", 2)
    assessor_names = collect_optional_names(assessors, "assessor", "A")
    participant_names = collect_optional_names(candidates, "participant", "P")

    start_minute = minutes_from_time(start_time)
    end_minute = minutes_from_time(end_time)
    slots_per_day = (end_minute - start_minute) // SLOT_MINUTES

    if slots_per_day <= 0:
        raise ValueError("The DC day is too short for a 5-minute schedule.")

    lunch_window_start = max(start_minute, LUNCH_WINDOW_START_MINUTE)
    lunch_window_end = min(end_minute, LUNCH_WINDOW_END_MINUTE)
    if lunch_window_end - lunch_window_start < LUNCH_MINUTES:
        raise ValueError("The DC day must include at least 30 minutes inside 12:00-16:00.")

    tools = []
    for index in range(1, tool_count + 1):
        print(f"\nTool {index}")
        name = input("Tool name: ").strip()
        while not name:
            print("Tool name cannot be empty.")
            name = input("Tool name: ").strip()

        if "bei" in name.lower():
            instruction_minutes = 0
            print("Instruction time set to 0 because tool name contains 'bei'.")
        else:
            instruction_minutes = read_non_negative_int("Instruction time (minutes): ")

        preparation_minutes = read_non_negative_int("Preparation time (minutes): ")
        execution_minutes = read_non_negative_int("Execution time (minutes): ")
        scoring_minutes = read_non_negative_int("Scoring time (minutes): ")

        instruction_slots, rounded_instruction_minutes = rounded_duration(instruction_minutes)
        preparation_slots, rounded_preparation_minutes = rounded_duration(preparation_minutes)
        execution_slots, rounded_execution_minutes = rounded_duration(execution_minutes)
        scoring_slots, rounded_scoring_minutes = rounded_duration(scoring_minutes)

        tools.append({
            "index": index - 1,
            "name": name,
            "instruction_slots": instruction_slots,
            "preparation_slots": preparation_slots,
            "execution_slots": execution_slots,
            "scoring_slots": scoring_slots,
            "instruction_minutes": rounded_instruction_minutes,
            "preparation_minutes": rounded_preparation_minutes,
            "execution_minutes": rounded_execution_minutes,
            "scoring_minutes": rounded_scoring_minutes,
        })

    context_slots, rounded_context_minutes = rounded_duration(context_minutes)
    if context_slots > slots_per_day:
        raise ValueError("Context setting duration is longer than one DC day.")

    return {
        "candidates": candidates,
        "assessors": assessors,
        "start_time": start_time,
        "end_time": end_time,
        "start_minute": start_minute,
        "end_minute": end_minute,
        "slots_per_day": slots_per_day,
        "context_slots": context_slots,
        "context_minutes": rounded_context_minutes,
        "tools": tools,
        "assessor_names": assessor_names,
        "participant_names": participant_names,
    }


def participant_code(number):
    return f"P{number}"


def assessor_code(index):
    return f"A{index + 1}"


def assessor_display_name(index, assessor_names):
    name = assessor_names[index].strip()
    return f"{assessor_code(index)} - {name}" if name else assessor_code(index)


def assessor_fill(index):
    return PatternFill("solid", fgColor=ASSESSOR_COLORS[index % len(ASSESSOR_COLORS)])


def lunch_fill():
    return PatternFill("solid", fgColor="F2F2F2")


def phase_sequence(tool):
    return [
        ("Instr", tool["instruction_slots"]),
        ("Prep", tool["preparation_slots"]),
        ("Exec", tool["execution_slots"]),
        ("Score", tool["scoring_slots"]),
    ]


def tool_total_slots(tool):
    return sum(duration for _, duration in phase_sequence(tool))


def phase_needs_assessor(phase_name):
    return phase_name in ("Exec", "Score")


def paired_assessor_index(index, assessor_count):
    if assessor_count % 2 == 0:
        return index + 1 if index % 2 == 0 else index - 1
    return (index + 1) % assessor_count


def create_groups(candidate_count, assessor_count):
    groups = []
    candidate_number = 1
    group_index = 0

    while candidate_number <= candidate_count:
        group_size = min(assessor_count, candidate_count - candidate_number + 1)
        members = list(range(candidate_number, candidate_number + group_size))
        primary_assessor = group_index % assessor_count

        groups.append({
            "name": f"G{group_index + 1}",
            "index": group_index,
            "members": members,
            "primary_assessor": primary_assessor,
            "secondary_assessor": paired_assessor_index(primary_assessor, assessor_count),
            "ready_slot": 0,
            "next_tool_index": 0,
            "tool_order": [],
            "lunch_days": set(),
        })

        candidate_number += group_size
        group_index += 1

    return groups


def rotate_tools_for_groups(groups, tools, rotations=None):
    for index, group in enumerate(groups):
        rotation = rotations[index] % len(tools) if rotations is not None else index % len(tools)
        group["tool_order"] = tools[rotation:] + tools[:rotation]


def participant_assessor_pair(participant, assessor_count):
    primary = (participant - 1) % assessor_count
    secondary = paired_assessor_index(primary, assessor_count)
    return primary, secondary


def assigned_assessor_for_participant(participant, tool, assessor_count):
    primary, secondary = participant_assessor_pair(participant, assessor_count)
    return primary if tool["index"] % 2 == 0 else secondary


def group_assessor_assignments(group, tool, inputs):
    return {
        participant: assigned_assessor_for_participant(participant, tool, inputs["assessors"])
        for participant in group["members"]
    }


def day_index(slot, inputs):
    return slot // inputs["slots_per_day"]


def local_slot(slot, inputs):
    return slot % inputs["slots_per_day"]


def day_start_slot(day, inputs):
    return day * inputs["slots_per_day"]


def day_end_slot(day, inputs):
    return day_start_slot(day, inputs) + inputs["slots_per_day"]


def minute_for_slot(slot, inputs):
    return inputs["start_minute"] + local_slot(slot, inputs) * SLOT_MINUTES


def time_text_from_minute(minute):
    return f"{minute // 60:02d}:{minute % 60:02d}"


def format_time(inputs, slot, include_day=False):
    label = time_text_from_minute(minute_for_slot(slot, inputs))
    return f"Day {day_index(slot, inputs) + 1} {label}" if include_day else label


def format_boundary_time(inputs, slot, include_day=False):
    if slot > 0 and slot % inputs["slots_per_day"] == 0:
        boundary_day = slot // inputs["slots_per_day"] - 1
        label = time_text_from_minute(inputs["end_minute"])
        return f"Day {boundary_day + 1} {label}" if include_day else label
    return format_time(inputs, slot, include_day)


def format_time_range(inputs, slot, include_day=False):
    start_label = time_text_from_minute(minute_for_slot(slot, inputs))
    end_label = time_text_from_minute(minute_for_slot(slot, inputs) + SLOT_MINUTES)
    if include_day:
        return f"Day {day_index(slot, inputs) + 1} {start_label}-{end_label}"
    return f"{start_label}-{end_label}"


def normalize_start_slot(slot, total_slots, inputs):
    if local_slot(slot, inputs) + total_slots > inputs["slots_per_day"]:
        return day_start_slot(day_index(slot, inputs) + 1, inputs)
    return slot


def lunch_window_text():
    return f"{time_text_from_minute(LUNCH_WINDOW_START_MINUTE)}-{time_text_from_minute(LUNCH_WINDOW_END_MINUTE)}"


def lunch_window_slot_range(day, inputs):
    window_start = max(inputs["start_minute"], LUNCH_WINDOW_START_MINUTE)
    window_end = min(inputs["end_minute"], LUNCH_WINDOW_END_MINUTE)

    if window_end - window_start < LUNCH_MINUTES:
        return None

    start_local = int(math.ceil((window_start - inputs["start_minute"]) / SLOT_MINUTES))
    end_local = int(math.floor((window_end - inputs["start_minute"]) / SLOT_MINUTES))

    earliest_start = day_start_slot(day, inputs) + start_local
    latest_start = day_start_slot(day, inputs) + end_local - lunch_slots()

    if earliest_start > latest_start:
        return None

    return earliest_start, latest_start


def lunch_target_slot(day, inputs):
    window = lunch_window_slot_range(day, inputs)
    if window is None:
        raise ValueError(f"No 30-minute lunch window exists on Day {day + 1}.")

    earliest_start, latest_start = window
    return round((earliest_start + latest_start) / 2)


def lunch_target_text(inputs):
    start = lunch_target_slot(0, inputs)
    end = start + lunch_slots()
    return f"{format_time(inputs, start)}-{format_boundary_time(inputs, end)}"


def lunch_start_candidates(day, inputs, earliest=None, latest=None):
    window = lunch_window_slot_range(day, inputs)
    if window is None:
        return []

    window_earliest, window_latest = window
    min_start = max(window_earliest, earliest if earliest is not None else window_earliest)
    max_start = min(window_latest, latest if latest is not None else window_latest)

    if min_start > max_start:
        return []

    target = lunch_target_slot(day, inputs)
    return sorted(range(min_start, max_start + 1), key=lambda slot: (abs(slot - target), slot))


def group_is_free_for_lunch(schedule, group, start_slot):
    for slot in range(start_slot, start_slot + lunch_slots()):
        for candidate in group["members"]:
            if schedule[candidate].get(slot):
                return False
    return True


def place_group_lunch(schedule, group, start_slot, inputs):
    day = day_index(start_slot, inputs)
    window = lunch_window_slot_range(day, inputs)

    if window is None:
        raise ValueError(f"No lunch window exists on Day {day + 1}.")

    earliest, latest = window
    if not earliest <= start_slot <= latest:
        raise ValueError(f"Lunch for {group['name']} must start inside {lunch_window_text()}.")

    if not group_is_free_for_lunch(schedule, group, start_slot):
        raise ValueError(f"Lunch for {group['name']} overlaps another activity.")

    for slot in range(start_slot, start_slot + lunch_slots()):
        for candidate in group["members"]:
            schedule[candidate][slot] = "Lunch"

    group["lunch_days"].add(day)
    group["ready_slot"] = max(group["ready_slot"], start_slot + lunch_slots())


def schedule_group_lunch_in_gap(schedule, group, gap_start, gap_end, inputs):
    if gap_start >= gap_end:
        return False

    scheduled = False
    first_day = day_index(gap_start, inputs)
    last_day = day_index(gap_end - 1, inputs)

    for day in range(first_day, last_day + 1):
        if day in group["lunch_days"]:
            continue

        earliest = max(gap_start, day_start_slot(day, inputs))
        latest = min(gap_end - lunch_slots(), day_end_slot(day, inputs) - lunch_slots())

        for lunch_start in lunch_start_candidates(day, inputs, earliest, latest):
            if group_is_free_for_lunch(schedule, group, lunch_start):
                place_group_lunch(schedule, group, lunch_start, inputs)
                scheduled = True
                break

    return scheduled


def should_lunch_before_tool(group, tool_start, tool_end, inputs):
    day = day_index(tool_start, inputs)

    if day in group["lunch_days"]:
        return False

    window = lunch_window_slot_range(day, inputs)
    if window is None:
        raise ValueError(f"No lunch window exists on Day {day + 1}.")

    _, latest_lunch_start = window
    target = lunch_target_slot(day, inputs)

    if tool_start >= target:
        return True

    if tool_end > latest_lunch_start:
        return True

    return False


def schedule_lunch_before_next_tool(schedule, group, tool_start, inputs):
    day = day_index(tool_start, inputs)
    earliest = max(group["ready_slot"], day_start_slot(day, inputs))

    for lunch_start in lunch_start_candidates(day, inputs, earliest, tool_start):
        if group_is_free_for_lunch(schedule, group, lunch_start):
            place_group_lunch(schedule, group, lunch_start, inputs)
            return True

    for lunch_start in lunch_start_candidates(day, inputs, earliest):
        if group_is_free_for_lunch(schedule, group, lunch_start):
            place_group_lunch(schedule, group, lunch_start, inputs)
            return True

    raise ValueError(f"Could not schedule required lunch for {group['name']} on Day {day + 1}.")


def schedule_missing_group_lunches(schedule, groups, inputs, total_days):
    for day in range(total_days):
        for group in groups:
            if day in group["lunch_days"]:
                continue

            for lunch_start in lunch_start_candidates(day, inputs):
                if group_is_free_for_lunch(schedule, group, lunch_start):
                    place_group_lunch(schedule, group, lunch_start, inputs)
                    break
            else:
                raise ValueError(f"Could not schedule lunch for {group['name']} on Day {day + 1}.")


def schedule_context(schedule, candidate_count, context_slots):
    for candidate in range(1, candidate_count + 1):
        for slot in range(context_slots):
            schedule[candidate][slot] = "Context Setting"


def block_feasible(start_slot, tool, assignments, exec_counts, score_counts, assessor_busy, inputs):
    if normalize_start_slot(start_slot, tool_total_slots(tool), inputs) != start_slot:
        return False

    current_slot = start_slot

    for phase_name, duration_slots in phase_sequence(tool):
        for offset in range(duration_slots):
            slot = current_slot + offset

            if phase_needs_assessor(phase_name):
                for assessor in assignments.values():
                    if assessor_busy[assessor].get(slot):
                        return False

            if phase_name == "Exec":
                if exec_counts.get(slot, 0) + len(assignments) > inputs["assessors"]:
                    return False
                if score_counts.get(slot, 0) > 0:
                    return False

            if phase_name == "Score":
                if exec_counts.get(slot, 0) > 0:
                    return False
                if score_counts.get(slot, 0) + len(assignments) > inputs["assessors"]:
                    return False

        current_slot += duration_slots

    return True


def apply_block(
    schedule,
    schedule_assessors,
    group,
    tool,
    assignments,
    start_slot,
    exec_counts,
    score_counts,
    assessor_busy,
    summary,
    assessor_mapping,
    participant_assessors,
    inputs,
):
    current_slot = start_slot
    block_start = start_slot

    for phase_name, duration_slots in phase_sequence(tool):
        for offset in range(duration_slots):
            slot = current_slot + offset
            value = f"{tool['name']} {phase_name}"

            for candidate in group["members"]:
                schedule[candidate][slot] = value
                if phase_needs_assessor(phase_name):
                    schedule_assessors[candidate][slot] = assignments[candidate]

            if phase_needs_assessor(phase_name):
                for candidate, assessor in assignments.items():
                    assessor_busy[assessor][slot] = (
                        f"{group['name']} {participant_code(candidate)} {tool['name']} {phase_name}"
                    )

            if phase_name == "Exec":
                exec_counts[slot] += len(assignments)
            elif phase_name == "Score":
                score_counts[slot] += len(assignments)

        current_slot += duration_slots

    for candidate in group["members"]:
        assessor = assignments[candidate]
        assessor_mapping[assessor][tool["index"]].append(candidate)
        participant_assessors[candidate].add(assessor)

    include_day = current_slot > inputs["slots_per_day"]
    assessor_text = ", ".join(
        f"{participant_code(candidate)}:{assessor_display_name(assignments[candidate], inputs['assessor_names'])}"
        for candidate in group["members"]
    )

    summary.append({
        "group": group["name"],
        "members": ", ".join(participant_code(member) for member in group["members"]),
        "tool": tool["name"],
        "assessor": assessor_text,
        "start": format_time(inputs, block_start, include_day),
        "end": format_boundary_time(inputs, current_slot, include_day),
    })

    group["ready_slot"] = current_slot
    group["next_tool_index"] += 1


def current_max_slot(schedule, default=0):
    max_slot = default
    for slots in schedule.values():
        if slots:
            max_slot = max(max_slot, max(slots) + 1)
    return max_slot


def total_days_for_slot_count(max_slot, inputs):
    return max(1, math.ceil(max_slot / inputs["slots_per_day"]))


def assessor_has_lunch(assessor_busy, assessor, day, inputs):
    return any(
        assessor_busy[assessor].get(slot) == "Lunch"
        for slot in range(day_start_slot(day, inputs), day_end_slot(day, inputs))
    )


def schedule_assessor_lunches(assessor_busy, inputs, total_days):
    assessor_lunches = []

    for day in range(total_days):
        for assessor in range(inputs["assessors"]):
            if assessor_has_lunch(assessor_busy, assessor, day, inputs):
                continue

            for lunch_start in lunch_start_candidates(day, inputs):
                lunch_end = lunch_start + lunch_slots()

                if all(not assessor_busy[assessor].get(slot) for slot in range(lunch_start, lunch_end)):
                    for slot in range(lunch_start, lunch_end):
                        assessor_busy[assessor][slot] = "Lunch"

                    assessor_lunches.append({
                        "day": day,
                        "assessor": assessor,
                        "start_slot": lunch_start,
                        "end_slot": lunch_end,
                    })
                    break
            else:
                raise ValueError(f"Could not schedule lunch for {assessor_code(assessor)} on Day {day + 1}.")

    return assessor_lunches


def build_schedule(inputs):
    schedule = defaultdict(dict)
    schedule_assessors = defaultdict(dict)
    exec_counts = defaultdict(int)
    score_counts = defaultdict(int)
    assessor_busy = defaultdict(dict)
    summary = []
    warnings = set()

    tools = inputs["tools"]
    assessor_mapping = {
        assessor_index: {tool["index"]: [] for tool in tools}
        for assessor_index in range(inputs["assessors"])
    }
    participant_assessors = defaultdict(set)

    schedule_context(schedule, inputs["candidates"], inputs["context_slots"])

    groups = create_groups(inputs["candidates"], inputs["assessors"])
    rotate_tools_for_groups(groups, tools, inputs.get("group_rotations"))

    for group in groups:
        group["ready_slot"] = inputs["context_slots"]

    for tool in tools:
        if tool_total_slots(tool) > inputs["slots_per_day"]:
            raise ValueError(f"{tool['name']} is longer than one DC day.")

    while any(group["next_tool_index"] < len(tools) for group in groups):
        groups.sort(key=lambda item: (item["ready_slot"], item["index"]))

        for group in groups:
            if group["next_tool_index"] >= len(tools):
                continue

            while group["next_tool_index"] < len(tools):
                tool = group["tool_order"][group["next_tool_index"]]
                assignments = group_assessor_assignments(group, tool, inputs)
                total_slots = tool_total_slots(tool)
                start_slot = group["ready_slot"]
                search_limit = start_slot + inputs["slots_per_day"] * 30

                while True:
                    start_slot = normalize_start_slot(start_slot, total_slots, inputs)

                    if start_slot > search_limit:
                        raise ValueError(f"Could not schedule {group['name']} for {tool['name']} within 30 DC days.")

                    if block_feasible(start_slot, tool, assignments, exec_counts, score_counts, assessor_busy, inputs):
                        break

                    start_slot += 1

                if group["ready_slot"] < start_slot:
                    schedule_group_lunch_in_gap(schedule, group, group["ready_slot"], start_slot, inputs)

                tool_end = start_slot + total_slots

                if should_lunch_before_tool(group, start_slot, tool_end, inputs):
                    schedule_lunch_before_next_tool(schedule, group, start_slot, inputs)
                    continue

                apply_block(
                    schedule,
                    schedule_assessors,
                    group,
                    tool,
                    assignments,
                    start_slot,
                    exec_counts,
                    score_counts,
                    assessor_busy,
                    summary,
                    assessor_mapping,
                    participant_assessors,
                    inputs,
                )
                break

    max_slot = current_max_slot(schedule, inputs["context_slots"])
    total_days = total_days_for_slot_count(max_slot, inputs)

    schedule_missing_group_lunches(schedule, groups, inputs, total_days)

    max_slot = current_max_slot(schedule, inputs["context_slots"])
    total_days = total_days_for_slot_count(max_slot, inputs)

    assessor_lunches = schedule_assessor_lunches(assessor_busy, inputs, total_days)

    return {
        "schedule": schedule,
        "schedule_assessors": schedule_assessors,
        "groups": groups,
        "summary": summary,
        "exec_counts": exec_counts,
        "score_counts": score_counts,
        "assessor_busy": assessor_busy,
        "assessor_mapping": assessor_mapping,
        "participant_assessors": participant_assessors,
        "assessor_lunches": assessor_lunches,
        "warnings": sorted(warnings),
        "max_slot": max_slot,
        "total_days": total_days,
    }


def quality_check_schedule(result, inputs):
    errors = []

    for slot in set(result["exec_counts"].keys()) | set(result["score_counts"].keys()):
        if result["exec_counts"][slot] > inputs["assessors"]:
            errors.append(f"Too many executions at {format_time_range(inputs, slot, True)}.")

        if result["score_counts"][slot] > inputs["assessors"]:
            errors.append(f"Too many scorings at {format_time_range(inputs, slot, True)}.")

        if result["exec_counts"][slot] and result["score_counts"][slot]:
            errors.append(f"Execution and scoring overlap at {format_time_range(inputs, slot, True)}.")

    for participant in range(1, inputs["candidates"] + 1):
        expected = min(2, inputs["assessors"])
        actual = len(result["participant_assessors"][participant])
        if actual != expected:
            errors.append(f"{participant_code(participant)} has {actual} assessor(s), expected {expected}.")

    for group in result["groups"]:
        for day in range(result["total_days"]):
            lunch_count = 0

            for slot in range(day_start_slot(day, inputs), day_end_slot(day, inputs)):
                values = {result["schedule"][candidate].get(slot) for candidate in group["members"]}
                if values == {"Lunch"}:
                    lunch_count += 1

            if lunch_count != lunch_slots():
                errors.append(
                    f"{group['name']} has {slots_to_minutes(lunch_count)} lunch minutes "
                    f"on Day {day + 1}, expected {LUNCH_MINUTES}."
                )

    for assessor in range(inputs["assessors"]):
        for day in range(result["total_days"]):
            lunch_count = sum(
                1
                for slot in range(day_start_slot(day, inputs), day_end_slot(day, inputs))
                if result["assessor_busy"][assessor].get(slot) == "Lunch"
            )

            if lunch_count != lunch_slots():
                errors.append(
                    f"{assessor_code(assessor)} has {slots_to_minutes(lunch_count)} lunch minutes "
                    f"on Day {day + 1}, expected {LUNCH_MINUTES}."
                )

    return errors


def validate_schedule(result, inputs):
    errors = quality_check_schedule(result, inputs)
    if errors:
        raise ValueError("Schedule failed quality checks:\n" + "\n".join(f"- {error}" for error in errors))


def write_schedule_sheet(workbook, inputs, result, sheet_title="Assessor Schedule", hide_score=False, participant_filter=None):
    ws = workbook.active if sheet_title == "Assessor Schedule" else workbook.create_sheet(sheet_title)
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False

    if participant_filter is not None:
        display_candidates = [participant_filter]
    else:
        display_candidates = list(range(1, inputs["candidates"] + 1))

    schedule_column_count = len(display_candidates) + 1
    time_repeat_column = schedule_column_count + 1
    mapping_start_column = schedule_column_count + 3
    mapping_column_count = len(inputs["tools"]) + 2
    mapping_end_column = mapping_start_column + mapping_column_count - 1
    used_max_column = mapping_end_column
    max_slot = result["max_slot"]
    include_day = result["total_days"] > 1

    participant_table_start_row = 3 + inputs["assessors"] + 2
    participant_table_end_row = participant_table_start_row + inputs["candidates"]
    used_max_row = max(max_slot + 3, participant_table_end_row)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=used_max_column)
    ws["A1"] = "DC Schedule"

    tool_names = ", ".join(tool["name"] for tool in inputs["tools"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=used_max_column)
    ws["A2"] = (
        f"Start: {inputs['start_time'].strftime('%H:%M')} | "
        f"End: {inputs['end_time'].strftime('%H:%M')} | "
        f"Candidates: {inputs['candidates']} | "
        f"Assessors: {inputs['assessors']} | "
        f"Lunch Window: {lunch_window_text()} | "
        f"Ideal Lunch: {lunch_target_text(inputs)} | "
        f"Tools: {tool_names}"
    )

    ws.cell(row=3, column=1, value="Time")
    for column, candidate in enumerate(display_candidates, start=2):
        ws.cell(row=3, column=column, value=participant_code(candidate))
    ws.cell(row=3, column=time_repeat_column, value="Time")

    for slot in range(max_slot):
        row = slot + 4
        time_label = format_time_range(inputs, slot, include_day)
        ws.cell(row=row, column=1, value=time_label)
        ws.cell(row=row, column=time_repeat_column, value=time_label)

        for column, candidate in enumerate(display_candidates, start=2):
            value = result["schedule"][candidate].get(slot, "")
            cell = ws.cell(row=row, column=column)

            if hide_score and isinstance(value, str) and value.endswith(" Score"):
                cell.value = ""
            else:
                cell.value = value

            if value == "Lunch":
                cell.fill = lunch_fill()
            elif isinstance(value, str) and value.endswith(" Score") and not hide_score:
                cell.fill = PatternFill("solid", fgColor="BFBFBF")
            elif isinstance(value, str) and value.endswith(" Exec") and slot in result["schedule_assessors"][candidate]:
                cell.fill = assessor_fill(result["schedule_assessors"][candidate][slot])

    if participant_filter is None:
        write_mapping_box(ws, inputs, result, 3, mapping_start_column)
        write_participant_table(ws, inputs, participant_table_start_row, mapping_start_column)

    format_schedule_sheet(ws, display_candidates, max_slot, used_max_column, used_max_row, time_repeat_column)

    if participant_filter is None:
        format_mapping_box(ws, inputs, 3, mapping_start_column)
        format_participant_table(ws, inputs, participant_table_start_row, mapping_start_column)

    merge_repeated_activities(ws, display_candidates, max_slot)


def write_mapping_box(ws, inputs, result, start_row, start_column):
    headers = ["Assessors", "Room Name"] + [tool["name"] for tool in inputs["tools"]]

    for offset, header in enumerate(headers):
        ws.cell(row=start_row, column=start_column + offset, value=header)

    for assessor_index in range(inputs["assessors"]):
        row = start_row + assessor_index + 1
        ws.cell(row=row, column=start_column, value=assessor_display_name(assessor_index, inputs["assessor_names"]))
        ws.cell(row=row, column=start_column + 1, value="")

        for tool in inputs["tools"]:
            participants = sorted(set(result["assessor_mapping"][assessor_index][tool["index"]]))
            ws.cell(
                row=row,
                column=start_column + 2 + tool["index"],
                value=",".join(str(participant) for participant in participants),
            )


def write_participant_table(ws, inputs, start_row, start_column):
    ws.cell(row=start_row, column=start_column, value="P.NO")
    ws.cell(row=start_row, column=start_column + 1, value="Participant Name")

    for participant in range(1, inputs["candidates"] + 1):
        row = start_row + participant
        ws.cell(row=row, column=start_column, value=participant_code(participant))
        ws.cell(row=row, column=start_column + 1, value=inputs["participant_names"][participant - 1])


def format_schedule_sheet(ws, display_candidates, max_slot, used_max_column, used_max_row, time_repeat_column):
    dark_blue = PatternFill("solid", fgColor="1F4E78")
    light_blue = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].fill = dark_blue
    ws["A1"].font = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"].font = Font(name=FONT_NAME, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    schedule_max_column = len(display_candidates) + 1

    for row in (1, 2):
        for column in range(1, used_max_column + 1):
            ws.cell(row=row, column=column).border = border

    for column in list(range(1, schedule_max_column + 1)) + [time_repeat_column]:
        cell = ws.cell(row=3, column=column)
        cell.fill = light_blue
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in range(4, max_slot + 4):
        for column in list(range(1, schedule_max_column + 1)) + [time_repeat_column]:
            cell = ws.cell(row=row, column=column)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if column in (1, time_repeat_column):
                cell.border = border

    ws.column_dimensions["A"].width = 20
    for column in range(2, schedule_max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18
    ws.column_dimensions[get_column_letter(time_repeat_column)].width = 20

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 22

    for row in range(4, used_max_row + 1):
        ws.row_dimensions[row].height = 24

    ws.freeze_panes = "A4"


def format_mapping_box(ws, inputs, start_row, start_column):
    header_fill = PatternFill("solid", fgColor="5B6770")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    column_count = len(inputs["tools"]) + 2

    for offset in range(column_count):
        cell = ws.cell(row=start_row, column=start_column + offset)
        cell.fill = header_fill
        cell.font = Font(name=FONT_NAME, color="FFFFFF", bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for assessor_index in range(inputs["assessors"]):
        row = start_row + assessor_index + 1
        for offset in range(column_count):
            cell = ws.cell(row=row, column=start_column + offset)
            cell.fill = assessor_fill(assessor_index)
            cell.font = Font(name=FONT_NAME, bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [26, 18] + [24] * len(inputs["tools"])
    for offset, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_column + offset)].width = width


def format_participant_table(ws, inputs, start_row, start_column):
    header_fill = PatternFill("solid", fgColor="5B6770")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for offset in range(2):
        cell = ws.cell(row=start_row, column=start_column + offset)
        cell.fill = header_fill
        cell.font = Font(name=FONT_NAME, color="FFFFFF", bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(start_row + 1, start_row + inputs["candidates"] + 1):
        for offset in range(2):
            cell = ws.cell(row=row, column=start_column + offset)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if offset == 0:
                cell.font = Font(name=FONT_NAME, bold=True)


def merge_repeated_activities(ws, display_candidates, max_slot):
    first_row = 4
    last_row = max_slot + 3
    thin = Side(style="thin", color="808080")
    activity_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column in range(2, len(display_candidates) + 2):
        merge_start = first_row
        current_value = ws.cell(row=merge_start, column=column).value

        for row in range(first_row + 1, last_row + 2):
            next_value = ws.cell(row=row, column=column).value if row <= last_row else None

            if next_value != current_value:
                if current_value not in (None, ""):
                    for border_row in range(merge_start, row):
                        ws.cell(row=border_row, column=column).border = activity_border

                    if row - merge_start > 1:
                        ws.merge_cells(
                            start_row=merge_start,
                            start_column=column,
                            end_row=row - 1,
                            end_column=column,
                        )

                    ws.cell(row=merge_start, column=column).alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )

                merge_start = row
                current_value = next_value


def write_summary_sheet(workbook, inputs, result):
    ws = workbook.create_sheet("Group Summary")
    ws.sheet_view.showGridLines = False
    headers = ["Group", "Members", "Tool", "Assessor", "Start", "End"]

    for column, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column, value=header)

    for row, item in enumerate(result["summary"], start=2):
        ws.cell(row=row, column=1, value=item["group"])
        ws.cell(row=row, column=2, value=item["members"])
        ws.cell(row=row, column=3, value=item["tool"])
        ws.cell(row=row, column=4, value=item["assessor"])
        ws.cell(row=row, column=5, value=item["start"])
        ws.cell(row=row, column=6, value=item["end"])

    format_table_sheet(ws, len(result["summary"]) + 1, len(headers))


def write_assessor_lunch_sheet(workbook, inputs, result):
    ws = workbook.create_sheet("Assessor Lunches")
    ws.sheet_view.showGridLines = False
    headers = ["Day", "Assessor", "Lunch Start", "Lunch End"]

    for column, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column, value=header)

    lunches = sorted(result["assessor_lunches"], key=lambda item: (item["day"], item["assessor"]))

    for row, item in enumerate(lunches, start=2):
        include_day = result["total_days"] > 1
        ws.cell(row=row, column=1, value=f"Day {item['day'] + 1}")
        ws.cell(row=row, column=2, value=assessor_display_name(item["assessor"], inputs["assessor_names"]))
        ws.cell(row=row, column=3, value=format_time(inputs, item["start_slot"], include_day))
        ws.cell(row=row, column=4, value=format_boundary_time(inputs, item["end_slot"], include_day))

    format_table_sheet(ws, len(lunches) + 1, len(headers))
    ws.column_dimensions["B"].width = 28


def format_table_sheet(ws, max_row, max_column):
    light_blue = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column in range(1, max_column + 1):
        cell = ws.cell(row=1, column=column)
        cell.fill = light_blue
        cell.font = Font(name=FONT_NAME, bold=True)

    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            cell = ws.cell(row=row, column=column)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column in range(1, max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 22

    ws.freeze_panes = "A2"


def apply_workbook_template(workbook):
    for ws in workbook.worksheets:
        ws.sheet_view.showGridLines = False

        for row in ws.iter_rows():
            for cell in row:
                new_font = copy(cell.font)
                new_font.name = FONT_NAME
                cell.font = new_font


def create_excel(inputs, result, output_file=OUTPUT_FILE):
    validate_schedule(result, inputs)
    workbook = Workbook()

    write_schedule_sheet(workbook, inputs, result, "Assessor Schedule", hide_score=False)
    write_schedule_sheet(workbook, inputs, result, "Participant Schedule", hide_score=True)

    for candidate in range(1, inputs["candidates"] + 1):
        write_schedule_sheet(
            workbook,
            inputs,
            result,
            sheet_title=participant_code(candidate),
            hide_score=True,
            participant_filter=candidate,
        )

    write_summary_sheet(workbook, inputs, result)
    write_assessor_lunch_sheet(workbook, inputs, result)
    apply_workbook_template(workbook)

    try:
        workbook.save(output_file)
    except PermissionError:
        raise PermissionError(
            f"\nCannot save '{output_file}' because the file is open in Excel.\n"
            "Please close it and run the scheduler again."
        )


def print_warnings(result):
    for warning in result["warnings"]:
        print(f"Warning: {warning}")


def reindex_tools(tools):
    return [{**tool, "index": index} for index, tool in enumerate(tools)]


def lunch_quality_score(result, inputs):
    score = 0

    for group in result["groups"]:
        for day in group["lunch_days"]:
            target = lunch_target_slot(day, inputs)
            starts = []

            for slot in range(day_start_slot(day, inputs), day_end_slot(day, inputs)):
                if all(result["schedule"][candidate].get(slot) == "Lunch" for candidate in group["members"]):
                    previous_is_lunch = (
                        slot > day_start_slot(day, inputs)
                        and all(
                            result["schedule"][candidate].get(slot - 1) == "Lunch"
                            for candidate in group["members"]
                        )
                    )

                    if not previous_is_lunch:
                        starts.append(slot)

            if starts:
                score += min(abs(start - target) for start in starts)

    for lunch in result["assessor_lunches"]:
        score += abs(lunch["start_slot"] - lunch_target_slot(lunch["day"], inputs))

    return score


def find_fastest_schedule(inputs):
    tools = inputs["tools"]
    n = len(tools)
    groups = create_groups(inputs["candidates"], inputs["assessors"])
    group_count = len(groups)

    other_rotation_combos = list(product(range(n), repeat=max(group_count - 1, 0)))
    total = math.factorial(n) * len(other_rotation_combos)

    if total > 1_000_000:
        print(f"Warning: {total:,} combinations to try. This may take several minutes.")
    else:
        print(f"Trying {total:,} combinations to find the fastest schedule...")

    best_inputs = None
    best_result = None
    best_max_slot = None
    best_lunch_score = None

    for perm_indices in permutations(range(n)):
        perm_tools = reindex_tools([tools[index] for index in perm_indices])

        for other_rots in other_rotation_combos:
            rotations = [0] + list(other_rots)
            perm_inputs = {**inputs, "tools": perm_tools, "group_rotations": rotations}

            try:
                result = build_schedule(perm_inputs)
                validate_schedule(result, perm_inputs)
            except ValueError:
                continue

            lunch_score = lunch_quality_score(result, perm_inputs)

            if (
                best_result is None
                or result["max_slot"] < best_max_slot
                or (result["max_slot"] == best_max_slot and lunch_score < best_lunch_score)
            ):
                best_inputs = perm_inputs
                best_result = result
                best_max_slot = result["max_slot"]
                best_lunch_score = lunch_score

    if best_result is None:
        raise ValueError("No valid schedule found for any combination.")

    print("\nFastest schedule found:")
    for group in best_result["groups"]:
        order = " -> ".join(tool["name"] for tool in group["tool_order"])
        print(f"  {group['name']}: {order}")

    return best_inputs, best_result


def main():
    inputs = collect_inputs()
    best_inputs, result = find_fastest_schedule(inputs)
    create_excel(best_inputs, result)
    print_warnings(result)
    print(f"\nSuccess: schedule saved as {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
